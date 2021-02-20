#!/usr/bin python3
#*-*coding:utf-8


import os
import sys
import openpyxl

def writeLmbench(path,name,sheetname,lb,x):
    '''
    横向写表方法用于多轮结果二维列表的写入
    '''
    wb = openpyxl.load_workbook("%s/%s" %(path,name))
    ws = wb[sheetname]
    y = 3
    for ac in lb:
        for st in ac:
            if type(st) == str:
                if "K" in st:
                    bs = st[0:-1]
                    ws.cell(x,y,value = float(bs))
                else:
                    ws.cell(x,y,value = float(st))
            else:
                ws.cell(x,y,value = float(st))
            x += 1
        x = x - len(ac)
        # print(x)
        y +=1
    # print(len(ac))
    wb.save("%s/%s" %(path,name))

    return x


def writeColumn(path,name,sheetname,lb,x):
    '''
    横向写表方法用于多轮结果二维列表的写入
    '''
    wb = openpyxl.load_workbook("%s/%s" %(path,name))
    ws = wb[sheetname]
    y = 3
    for ac in lb:
        for st in ac:
            ws.cell(x,y,value = float(st))
            x += 1
        x = x - len(ac)
        y +=1
    wb.save("%s/%s" %(path,name))

    return x

# writeColumn()

def writeLine(path,name,sheetname,lb,x):
    '''
    横向写表方法用于多轮结果一维列表的写入
    '''
    wb = openpyxl.load_workbook("%s/%s" %(path,name))
    # print(path,name,sheetname)
    ws = wb[sheetname]
    
    y = 3
    for st in lb:
        ws.cell(x,y,value = st)
        y += 1

    # print(lb)
    # print("="*100)
    wb.save("%s/%s" %(path,name))

    return x
