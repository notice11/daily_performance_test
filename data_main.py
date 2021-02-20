#!/usr/bin/python3
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
import numpy as np
import xlrd
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
from data_tools import *
from data_tools import Data_tools

class Data_main():

    def __init__(self,path):
        self.path = path

        #写表列数 y 初始值 5
        y = 5
        self.y = y

        #增加cpu信息确定测试平台
        framework = Data_tools.execCmd("cat %s/cpuname.conf |grep 'Architecture:' |awk '{print $2}'" %(path))
        framework = ",".join(framework)
        print ("cpu架构: %s" %framework)
        if "mips64" == framework:
            mipsname = Data_tools.execCmd("cat %s/cpuname.conf |grep 'Model name:' |awk '{print $5}'" %(path))
            mipsname = ",".join(mipsname)
            if mipsname == "(Loongson-3A3000)":
                cpuname = "3A3000"
            else:
                cpuname = "3A4000"
        else:
            model = Data_tools.execCmd("cat %s/cpuname.conf |grep 'Vendor ID:' |awk '{print $3}'" %(path))
            model = ",".join(model)
            if model == "GenuineIntel":
                cpuname = "Intel"
            elif model == "CentaurHauls":
                cpuname = "zhaoxin"
            elif model == "0x70":
                cpuname = "FT"
            elif model == "0x48":
                cpuname = "kunpeng"
            else:
                print ("未知的cpu架构,测试数据默认会写入intel 表内")
                cpuname = "Intel"

        print ("测试平台: %s" %cpuname)
        self.sheetname = cpuname

        #获得脚本所在的绝对路径
        folder = os.path.dirname(os.path.abspath(__file__))
        print (folder)
        self.folder = folder  
        #拷贝套件内的文件到测试数据所以在的目录
        os.system("cp %s/report/result.xlsx  %s" %(self.folder,self.path))
        #遍历原始表格删除当前平台以外的其他表格信息
        wb = openpyxl.load_workbook("%s/result.xlsx" %self.path)
        ws = wb[self.sheetname]
        for a in wb:
            if a is ws:
                continue
            else:
                wb.remove(a)
        wb.save("%s/result.xlsx" %self.path)
         


    #写入表格的方法,获取表格,根据表格的 坐标写入数据到表格
    def xlsx(self,dataAvg,x):
        wb = openpyxl.load_workbook("%s/result.xlsx" %(self.path))
        ws = wb[self.sheetname]
        for st in dataAvg:
            ws.cell(x,self.y,value = float('%.2f' %st))
            x += 1
        
        wb.save("%s/result.xlsx" %self.path)
        return x

    #写入glxgears数据到表格
    def gs_xl(self):
        gs = Data_tools(self.path).glxgears()
        self.xlsx(gs,14)
    
    #写入glmark2数据到表格
    def gk_xl(self):
        gk = Data_tools(self.path).glmark2()
        self.xlsx(gk,15)


    #写入unixebnch2D数据到表格
    def un2d_xl(self):
        un2d = Data_tools(self.path).unixebnch2D()
        self.xlsx(un2d,16)

    #写入unixbench数据到表格
    def unix_xl(self):
        unix_one,unix_to = Data_tools(self.path).unixbench()
        x = self.xlsx(unix_one,18)
        self.xlsx(unix_to,x+1)

    #写入lmbench数据到表格
    def lm_xl(self):
        lm_1,lm_2,lm_3,lm_4,lm_5,lm_6,lm_7,lm_8,lm_9 = Data_tools(self.path).lmbench()
        x = self.xlsx(lm_1[1:],46)
        x = self.xlsx(lm_2,x+1)
        x = self.xlsx(lm_3,x+7)
        x = self.xlsx(lm_4,x+1)
        x = self.xlsx(lm_5,x+1)
        x = self.xlsx(lm_6,x+1)
        x = self.xlsx(lm_7,x+1)
        x = self.xlsx(lm_8,x+1)
        self.xlsx(lm_9,x+1)

    #写入iozone数据到表格
    def io_xl(self):
        max,mem,min = Data_tools(self.path).iozone()
        x = self.xlsx(max,121)
        x = self.xlsx(mem,x)
        self.xlsx(min,x)

    #写入stream数据到表格
    def sm_xl(self):
        sm_one,sm_to = Data_tools(self.path).stream()
        x = self.xlsx(sm_one,160)
        self.xlsx(sm_to,x)


    # #写入10G文件拷贝数据到表格
    def file_10(self):
        cp_10 = Data_tools(self.path).file_cp()
        self.xlsx(cp_10,174)

    # #写入10G文件拷贝数据到表格
    def fio(self):
        _512,_512i,_1,_1i = Data_tools(self.path).fio()
        x = self.xlsx(_512,180)
        x = self.xlsx(_512i,x)
        x = self.xlsx(_1,x)
        self.xlsx(_1i,x)

    # #邮件推送


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print('\n警告：运行程序需要输入测试数据所在的目录 \n请以如下格式重新运行程序:\n\tpython3 data_main.py /home/uos/logs \n或者 \n\tpython3 data_main.py /home/uos/fio')
        exit()
    #测试数据所在的目录
    path = sys.argv[1]
    print (path)
    da = Data_main(path)
    da.y = 5
    filename = os.listdir(path)
    if "iozone" in ",".join(filename) or "lmbench" in ",".join(filename) or "unixbench" in ",".join(filename):
        try:
            da.gs_xl()
        except:
            print ("da.gs_xl() 方法出现异常")

        try:
            da.gk_xl()
        except:
            print ("da.gk_xl() 方法出现异常")

        try:
            da.un2d_xl()
        except:
            print ("da.un2d_xl() 方法出现异常")

        try:
            da.lm_xl()
        except:
            print ("da.lm_xl() 方法出现异常")

        try:
            da.unix_xl()
        except:
            print ("da.unix_xl() 方法出现异常")

        try:    
            da.io_xl()
        except:
            print ("da.io_xl() 方法出现异常")

        try:
            da.sm_xl()
        except:
            print ("da.sm_xl() 方法出现异常")

        try:
            da.file_10()
        except:
            print ("da.file_10() 方法出现异常")

    elif "512B_" in ",".join(filename):
        try:
            da.fio()
        except:
            print ("da.fio() 方法出现异常")

    print (" \n 数据采集完成 \n 采集结果保存在: \n \t %sresult.xlsx 文件中\n " %path)
    

