#!/usr/bin python3
# -*- coding: utf-8 -*-

import os
import sys
import numpy as np
import openpyxl


class Unixbench:

    def __init__(self,path):
        self.path = path
        self.filename = os.listdir(self.path)

        #增加cpu信息确定测试平台
        framework = self.dcmd("cat %s/cpuname.conf |grep 'Architecture:' |awk '{print $2}'" %(path))
        framework = ",".join(framework)
        print ("cpu架构: %s" %framework)
        if "mips64" == framework:
            mipsname = self.dcmd("cat %s/cpuname.conf |grep 'Model name:' |awk '{print $5}'" %(path))
            mipsname = ",".join(mipsname)
            if mipsname == "(Loongson-3A3000)":
                cpuname = "3A3000"
            else:
                cpuname = "3A4000"
        else:
            model = self.dcmd("cat %s/cpuname.conf |grep 'Vendor ID:' |awk '{print $3}'" %(path))
            model = ",".join(model)
            if model == "GenuineIntel":
                cpuname = "Intel"
            elif model == "CentaurHauls":
                cpuname = "zhaoxin"
            elif model == "0x70":
                cpuname = "FT"
            elif model == "0x48":
                cpuname = "kunpeng"
            else:
                print ("未知的cpu架构,测试数据默认会写入intel 表内")
                cpuname = "Intel"

        print ("测试平台: %s" %cpuname)
        self.sheetname = cpuname

        #获得脚本所在的绝对路径
        folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        print (folder)
        self.folder = folder
        #拷贝套件内的文件到测试数据所以在的目录
        os.system("cp %s/report/fluctuation.xlsx %s/" %(self.folder,self.path))
        #遍历原始表格删除当前平台以外的其他表格信息
        wb = openpyxl.load_workbook("%s/fluctuation.xlsx" %self.path)
        ws = wb[self.sheetname]
        for a in wb:
            if a is ws:
                continue
            else:
                wb.remove(a)
        wb.save("%s/fluctuation.xlsx" %self.path)

    def dcmd(self,cmd):  
        r = os.popen(cmd)  
        text = str(r.read())
        log = text.rstrip('\n')
        log = log.split("  ")
        log = [x for x in log if x!=''] 
        r.close()  
        return log
    
    def xlsx(self,lb,x):
        wb = openpyxl.load_workbook("%s/fluctuation.xlsx" %(self.path))
        ws = wb[self.sheetname]
        y = 4
        for ac in lb:
            x = 11
            for st in ac:
                ws.cell(x,y,value = st)
                x += 1
            y +=1
        wb.save("%s/fluctuation.xlsx" %self.path)
        return x
    
    def xlsx1(self,lb,x):
        wb = openpyxl.load_workbook("%s/fluctuation.xlsx" %(self.path))
        ws = wb[self.sheetname]
        y = 4
        for ac in lb:
            x = 25
            for st in ac:
                ws.cell(x,y,value = st)
                x += 1
            y +=1
        wb.save("%s/fluctuation.xlsx" %self.path)
        return x

    def _xlsx(self,lb,x):
        wb = openpyxl.load_workbook("%s/fluctuation.xlsx" %(self.path))
        ws = wb[self.sheetname]
        y = 4
        for st in lb:
            ws.cell(x,y,value = (st))
            y += 1
        wb.save("%s/fluctuation.xlsx" %self.path)
        return x


    def _unix(self):
        _no1 = ""
        _nos = ""
        for a in self.filename:
            if "unixbench_Ncores.log" == a:
                _nos = a
            elif "unixbench_1core.log" == a:
                _no1 = a
        def readunix(name):
            individual = self.dcmd("cat  %s/%s |grep  -A 12 'System Benchmarks Index Values'" %(self.path,name))
            total_ = self.dcmd("cat  %s/%s |grep  'System Benchmarks Index Score' " %(self.path,name))
            total_ = (",".join(total_).split("\n"))
            # print (total_)
            _total = []
            for a in total_:
                _total.append(float(a.split(",")[1]))

            _individual = []
            for a in individual[2::3]:
                if " RESULT" == a:
                    continue
                else:
                    _individual.append(float(a))
            
            core = np.array(_individual).reshape(int(len(_individual)/12),12)
 
            # print (_individual[])
            # print(len(_total))
            # _individual.append(_total)
            return core,_total

        try:
            _individual,_individual_1 = readunix(_no1)
            _total,_total_1 = readunix(_nos)
        except:
            print("数据读取存在错误------>请检查")

        print("\n单核心参数:\n子项:\n%s\n总分:\n%s\n" %(_individual,_individual_1))
        print("\n多核核心参数:\n子项:\n%s\n总分:\n%s\n" %(_total,_total_1))

        x = self.xlsx(_individual,11)
        x = self._xlsx(_individual_1,x)
        x = self.xlsx1(_total,x+1)
        self._xlsx(_total_1,x)



if __name__ == "__main__":
    if len(sys.argv) < 2:
        print('\n警告：运行程序需要输入测试数据所在的目录绝对路径 \n请以如下格式重新运行程序:\n\tpython3 unixbench.py /home/uos/logs \n或者 \n\tpython3 unixbench.py /home/uos/fio')
        exit()
    #测试数据所在的目录
    path = sys.argv[1]
    un = Unixbench(path)
    un._unix()